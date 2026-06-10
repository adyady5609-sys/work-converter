import streamlit as st
import openpyxl
import io
from openpyxl.styles import Font, Alignment, PatternFill

st.set_page_config(page_title="근무 데이터 통합 변환기", page_icon="📊", layout="centered")

st.title("📊 근무 데이터 통합 변환기")
st.markdown("**클로드_연습.xlsm** 파일을 업로드하면 통합문서를 자동으로 만들어드립니다.")

SHEET_CONFIG = {
    '뉴뉴야간': {
        'nameRow': 1, 'colStep': 5,
        'rows': {'총근무': 42, '추가근무': 43, '야간근무': 44, '휴일근무': 45, '주휴수당': 46, '합계': 47}
    },
    'DEFAULT': {
        'nameRow': 2, 'colStep': 4,
        'rows': {'총근무': 43, '추가근무': 44, '야간근무': 45, '휴일근무': 46, '주휴수당': 47, '합계': 48}
    }
}
SKIP_SHEETS = ['중도퇴사자']
LABELS = ['총 근무시간', '추가근무시간', '야간근무시간', '휴일근무', '주휴수당', '합계']
FIELDS = ['총근무', '추가근무', '야간근무', '휴일근무', '주휴수당', '합계']

uploaded = st.file_uploader("파일 선택 (.xlsm 또는 .xlsx)", type=['xlsm', 'xlsx'])

if uploaded:
    st.success(f"✅ {uploaded.name} 업로드 완료!")

    if st.button("통합문서 만들기", type="primary", use_container_width=True):
        with st.spinner("변환 중..."):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
                employees = []
                sheet_count = 0

                progress = st.progress(0)
                sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]

                for i, sheet_name in enumerate(sheets):
                    ws = wb[sheet_name]
                    config = SHEET_CONFIG.get(sheet_name, SHEET_CONFIG['DEFAULT'])
                    col = 2
                    while col <= ws.max_column:
                        name_cell = ws.cell(row=config['nameRow'], column=col).value
                        if name_cell and isinstance(name_cell, str) and name_cell.strip():
                            val_col = col + 3
                            emp = {'이름': name_cell.strip(), '시트': sheet_name}
                            for field in FIELDS:
                                v = ws.cell(row=config['rows'][field], column=val_col).value
                                emp[field] = v if isinstance(v, (int, float)) else 0
                            employees.append(emp)
                        col += config['colStep']
                    sheet_count += 1
                    progress.progress((i + 1) / len(sheets))

                # 결과 엑셀 생성
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = 'Sheet1'
                ws_out.column_dimensions['A'].width = 15
                for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
                    ws_out.column_dimensions[col_letter].width = 12

                header_fill = PatternFill('solid', start_color='F2F2F2')
                row = 1
                for emp in employees:
                    for i, label in enumerate(LABELS):
                        cell = ws_out.cell(row=row, column=i + 2)
                        cell.value = label
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    row += 1
                    ws_out.cell(row=row, column=1).value = emp['이름']
                    ws_out.cell(row=row, column=1).font = Font(bold=True)
                    for i, field in enumerate(FIELDS):
                        val = emp[field]
                        cell = ws_out.cell(row=row, column=i + 2)
                        cell.value = val if val else None
                        cell.alignment = Alignment(horizontal='right')
                    row += 1

                out = io.BytesIO()
                wb_out.save(out)
                out.seek(0)

                night_workers = sum(1 for e in employees if e['야간근무'] > 0)

                col1, col2, col3 = st.columns(3)
                col1.metric("총 직원", f"{len(employees)}명")
                col2.metric("처리 시트", f"{sheet_count}개")
                col3.metric("야간근무자", f"{night_workers}명")

                st.download_button(
                    label="📥 통합문서 다운로드",
                    data=out,
                    file_name="통합문서_자동.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )

            except Exception as e:
                st.error(f"오류 발생: {str(e)}")
